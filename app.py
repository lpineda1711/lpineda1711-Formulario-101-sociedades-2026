import streamlit as st
from openpyxl import load_workbook
from copy import copy
from io import BytesIO

st.set_page_config(layout="wide")
st.title("Generador Final F101 Vinculado con BC")

uploaded_f101 = st.file_uploader(
    "Sube el F101 (.xlsx)",
    type=["xlsx"]
)

if uploaded_f101:

    # ===============================
    # CARGAR PLANTILLA BASE
    # ===============================
    plantilla_path = "CT 2024 f101 ARCTURUS NUEVO.xlsx"
    wb_base = load_workbook(plantilla_path)

    # Eliminar F101 anterior
    if "F101" in wb_base.sheetnames:
        wb_base.remove(wb_base["F101"])

    # ===============================
    # CARGAR F101 SUBIDO
    # ===============================
    wb_nuevo = load_workbook(uploaded_f101)
    hoja_origen = wb_nuevo.active

    hoja_destino = wb_base.create_sheet("F101")

    # ===============================
    # COPIAR F101 COMPLETO CON ESTILOS
    # ===============================
    for row in hoja_origen.iter_rows():
        for cell in row:
            new_cell = hoja_destino[cell.coordinate]
            new_cell.value = cell.value

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)

    # Copiar celdas combinadas
    for merged in hoja_origen.merged_cells.ranges:
        hoja_destino.merge_cells(str(merged))

    # Copiar ancho de columnas
    for col in hoja_origen.column_dimensions:
        hoja_destino.column_dimensions[col].width = hoja_origen.column_dimensions[col].width

    # ===============================
    # VINCULAR F101 CON BC (FORMULAS)
    # ===============================
    if "BC" in wb_base.sheetnames:

        hoja_bc = wb_base["BC"]
        merged_ranges = hoja_destino.merged_cells.ranges

        fila_inicio = 1
        fila_fin = hoja_destino.max_row

        for fila in range(fila_inicio, fila_fin + 1):
            codigo_cell = hoja_destino[f"N{fila}"]
            destino_cell = hoja_destino[f"O{fila}"]

            # Saltar filas sin código
            if codigo_cell.value in (None, ""):
                continue

            # Verificar si la celda O está combinada
            escribir = True
            for merged in merged_ranges:
                if destino_cell.coordinate in merged:
                    # Solo permitir escribir en la celda superior izquierda
                    if destino_cell.coordinate != merged.start_cell.coordinate:
                        escribir = False
                    break

            if not escribir:
                continue

            destino_cell.value = (
                f'=SUMAR.SI(BC!$E:$E;F101!N{fila};BC!$D:$D)'
            )

    # ===============================
    # REORDENAR HOJAS (BC PRIMERO)
    # ===============================
    wb_base._sheets.sort(key=lambda ws: ws.title != "BC")

    # ===============================
    # EXPORTAR ARCHIVO FINAL
    # ===============================
    output = BytesIO()
    wb_base.save(output)
    output.seek(0)

    st.success("F101 pegado correctamente y vinculado con BC mediante fórmulas")

    st.download_button(
        "Descargar archivo final",
        data=output,
        file_name="ARCHIVO_FINAL_VINCULADO.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
